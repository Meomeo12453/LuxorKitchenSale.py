import streamlit as st
from PIL import Image
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import plotly.express as px
import colorsys
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
import random
import base64
import uuid

# ========== LOGO & GIAO DIỆN =============
st.set_page_config(page_title="Sales Dashboard MiniApp", layout="wide")

for _ in range(4):
    st.write("")

st.markdown("""
    <style>
    .block-container {padding-top:0.7rem; max-width:100vw !important;}
    .stApp {background: #F7F8FA;}
    img { border-radius: 0 !important; }
    h1, h2, h3 { font-size: 1.18rem !important; font-weight:600; }
    </style>
""", unsafe_allow_html=True)

LOGO_PATHS = [
    "logo-daba.png",
    "ef5ac011-857d-4b32-bd70-ef9ac3817106.png"
]
logo = None
for path in LOGO_PATHS:
    if os.path.exists(path):
        logo = Image.open(path)
        break

if logo is not None:
    desired_height = 36
    w, h = logo.size
    new_width = int((w / h) * desired_height)
    logo_resized = logo.resize((new_width, desired_height))
    buffered = BytesIO()
    logo_resized.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode()
    st.markdown(
        f"""
        <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;width:100%;padding-top:4px;padding-bottom:0;">
            <img src="data:image/png;base64,{img_str}" 
                 width="{new_width}" height="{desired_height}" style="display:block;margin:auto;" />
            <div style="height:5px;"></div>
        </div>
        """,
        unsafe_allow_html=True
    )
else:
    st.warning("⚠️ Không tìm thấy logo! Vui lòng kiểm tra lại tên file/logo trong thư mục app.")

st.markdown(
    "<div style='text-align:center;font-size:20px;color:#1570af;font-weight:600;'>BẢNG TÍNH HOA HỒNG CÔNG TY TNHH DABA SAIGON</div>",
    unsafe_allow_html=True)
st.markdown(
    "<div style='text-align:center;font-size:14px;color:#555;'>Hotline 0909.625.808 Địa chỉ: Lầu 9, Pearl Plaza, 561A Điện Biên Phủ, P.25, Q. Bình Thạnh, TP.HCM</div>",
    unsafe_allow_html=True)
st.markdown("<hr style='margin:10px 0 20px 0;border:1px solid #EEE;'>", unsafe_allow_html=True)

# ========== CONTROL ==========
st.markdown("### 🔎 Tùy chọn phân tích")
col1, col2 = st.columns([2, 1])
with col1:
    chart_type = st.radio(
        "Chọn loại biểu đồ:",
        ["Biểu đồ cột chồng", "Sơ đồ Sunburst", "Biểu đồ Pareto", "Biểu đồ tròn (Pie)"],
        horizontal=True
    )
with col2:
    filter_nganh = st.multiselect("Lọc theo nhóm khách hàng:", ["Catalyst", "Visionary", "Trailblazer"], default=[])

st.markdown("<hr style='margin:10px 0 20px 0;border:1px solid #EEE;'>", unsafe_allow_html=True)

# ======= MULTI FILE UPLOAD =======
st.markdown("### 1. Tải lên tối đa 10 file Excel (.xlsx)")
uploaded_files = st.file_uploader(
    "**Chọn nhiều file hoặc kéo thả nhiều file Excel**",
    type="xlsx",
    accept_multiple_files=True,
    help="Chỉ nhận Excel, <200MB mỗi file. Các file phải cùng cấu trúc cột."
)
if not uploaded_files:
    st.info("💡 Hãy upload 1 hoặc nhiều file Excel mẫu để bắt đầu sử dụng Dashboard.")
    with st.expander("📋 Xem hướng dẫn & file mẫu", expanded=False):
        st.markdown(
            "- Chọn hoặc kéo thả **1–10 file Excel**.\n"
            "- File cần các cột: **Mã khách hàng, Tên khách hàng, Nhóm khách hàng, Tổng bán trừ trả hàng, Ghi chú**.\n"
            "- Nếu lỗi, kiểm tra lại tiêu đề cột trong file Excel."
        )
    st.stop()

# ===== GỘP & LÀM SẠCH DỮ LIỆU =====
dfs = []
for f in uploaded_files[:10]:
    dft = pd.read_excel(f)
    dfs.append(dft)
df = pd.concat(dfs, ignore_index=True)

if any(df['Tên khách hàng'].astype(str).str.contains('[^\x00-\x7F]', na=False)):
    st.info("ℹ️ File có chứa ký tự đặc biệt hoặc tiếng Việt. Nếu bị lỗi font khi mở file Excel, hãy lưu lại bằng Excel phiên bản quốc tế hoặc UTF-8.")

required_cols = ['Mã khách hàng','Nhóm khách hàng','Tổng bán trừ trả hàng','Ghi chú','Tên khách hàng']
missing_cols = [col for col in required_cols if col not in df.columns]
if missing_cols:
    all_cols_lower = [c.lower().replace(" ", "").replace("_", "") for c in df.columns]
    for req in required_cols:
        if req.lower().replace(" ", "").replace("_", "") not in all_cols_lower:
            st.error(f"Thiếu cột '{req}' trong file Excel. Hãy kiểm tra lại tiêu đề cột (có thể bị thiếu dấu hoặc sai chính tả)!")
    st.stop()

n_trung = len(df) - df['Mã khách hàng'].nunique()
if n_trung > 0:
    st.warning(f"⚠️ Có {n_trung} dòng dữ liệu bị trùng mã khách hàng và đã bị loại bỏ. Vui lòng kiểm tra file gốc.")

df['Mã khách hàng'] = df['Mã khách hàng'].astype(str).str.strip()
df['Ghi chú'] = df['Ghi chú'].astype(str).str.strip()
df['Ghi chú'] = df['Ghi chú'].replace({'None': None, 'nan': None, 'NaN': None, '': None})
df['Tổng bán trừ trả hàng'] = pd.to_numeric(df['Tổng bán trừ trả hàng'], errors='coerce').fillna(0)
df = df.drop_duplicates(subset=['Mã khách hàng'], keep='first')

if (df['Tổng bán trừ trả hàng'] == 0).all():
    st.warning("⚠️ Tất cả doanh số đều bằng 0. Kiểm tra lại dữ liệu đầu vào!")
if df['Tổng bán trừ trả hàng'].isnull().any():
    st.warning("⚠️ Có dòng bị thiếu doanh số. Đã tự động điền 0 nhưng nên kiểm tra lại file gốc.")

null_kh = df['Mã khách hàng'].isnull().sum()
if null_kh > 0:
    st.warning(f"⚠️ Có {null_kh} dòng thiếu mã khách hàng! Đã loại bỏ khỏi kết quả.")

if df['Ghi chú'].str.contains(',|;|/|\\| ').any():
    st.warning("⚠️ Một số dòng 'Ghi chú cấp bậc' chứa nhiều mã hoặc ký tự phân cách (dấu phẩy, chấm phẩy, khoảng trắng, ...). Ứng dụng chỉ lấy mã đầu tiên.")

all_codes = set(df['Mã khách hàng'])

def get_parent_id(x):
    if pd.isnull(x) or x is None:
        return None
    x = str(x).strip()
    return x if x in all_codes else None
df['parent_id'] = df['Ghi chú'].apply(get_parent_id)

invalid_parents = df[(df['Ghi chú'].notnull()) & (~df['Ghi chú'].isin(all_codes))]
if len(invalid_parents) > 0:
    st.warning(f"⚠️ Có {len(invalid_parents)} dòng có 'Ghi chú phân cấp' không khớp mã khách hàng nào. Các dòng này sẽ không được tính phân cấp.")

parent_map = {}
for idx, row in df.iterrows():
    pid = row['parent_id']
    code = row['Mã khách hàng']
    if pd.notnull(pid) and pid is not None:
        parent_map.setdefault(pid, []).append(code)

def detect_cycles(parent_map):
    cycles = []
    def visit(node, visited):
        if node in visited:
            return True
        visited.add(node)
        for child in parent_map.get(node, []):
            if visit(child, visited):
                cycles.append((node, child))
        visited.remove(node)
        return False
    for k in parent_map.keys():
        visit(k, set())
    return set(cycles)
cycles = detect_cycles(parent_map)
if cycles:
    st.warning(f"⚠️ Chú ý các cấp bậc quản lý đang có nhiều thuộc cấp")

def get_all_descendants(code, parent_map, visited=None):
    if visited is None:
        visited = set()
    result = []
    children = parent_map.get(code, [])
    for child in children:
        if child not in visited:
            visited.add(child)
            result.append(child)
            result.extend(get_all_descendants(child, parent_map, visited))
    return result

desc_counts = []
ds_he_thong = []
for idx, row in df.iterrows():
    code = row['Mã khách hàng']
    descendants = get_all_descendants(code, parent_map, visited=set([code]))
    desc_counts.append(len(descendants))
    doanhso = df[df['Mã khách hàng'].isin(descendants)]['Tổng bán trừ trả hàng'].sum() if descendants else 0
    ds_he_thong.append(doanhso)
df['Số cấp dưới'] = desc_counts
df['Doanh số hệ thống'] = ds_he_thong

# Tính comm và override_comm
network = {
    'Catalyst':     {'comm_rate': 0.35, 'override_rate': 0.00},
    'Visionary':    {'comm_rate': 0.40, 'override_rate': 0.05},
    'Trailblazer':  {'comm_rate': 0.40, 'override_rate': 0.05},
}
df['comm_rate']     = df['Nhóm khách hàng'].map(lambda r: network.get(r, {}).get('comm_rate', 0))
df['override_rate'] = df['Nhóm khách hàng'].map(lambda r: network.get(r, {}).get('override_rate', 0))
df['override_comm'] = df['Doanh số hệ thống'] * df['override_rate']

# === BỔ SUNG: HỆ THỐNG VƯỢT CẤP (Trailblazer và các Catalyst trực thuộc) ===
trailblazer_codes = df[df['Nhóm khách hàng'] == 'Trailblazer']['Mã khách hàng'].astype(str)
catalyst_children = df[(df['Nhóm khách hàng'] == 'Catalyst') & (df['parent_id'].notnull())]
catalyst_children = catalyst_children[catalyst_children['parent_id'].isin(trailblazer_codes)]
vuot_cap_ds = catalyst_children.groupby('parent_id')['Tổng bán trừ trả hàng'].sum()
vuot_cap_hh = vuot_cap_ds * 0.10

# Gắn cột doanh số vượt cấp và hoa hồng vượt cấp vào đúng Trailblazer
df['Doanh số vượt cấp'] = df['Mã khách hàng'].astype(str).map(vuot_cap_ds).fillna(0)
df['Hoa hồng vượt cấp'] = df['Mã khách hàng'].astype(str).map(vuot_cap_hh).fillna(0)
# Đánh dấu Catalyst thuộc hệ thống vượt cấp của ai (Trailblazer nào)
catalyst_sys_map = catalyst_children.set_index('Mã khách hàng')['parent_id'].to_dict()
df['vuot_cap_trailblazer'] = df['Mã khách hàng'].map(catalyst_sys_map)

# ==== Thay đổi thứ tự cột xuất Excel: đặt "Doanh số vượt cấp" trước "Hoa hồng vượt cấp" ====
cols = list(df.columns)
if 'Hoa hồng vượt cấp' in cols and 'Doanh số vượt cấp' in cols:
    cols.remove('Doanh số vượt cấp')
    idx_hhvc = cols.index('Hoa hồng vượt cấp')
    cols.insert(idx_hhvc, 'Doanh số vượt cấp')
df = df[cols]

if filter_nganh:
    df = df[df['Nhóm khách hàng'].isin(filter_nganh)]

if len(df) > 1000:
    st.warning("⚠️ Dữ liệu quá nhiều khách hàng. Một số biểu đồ có thể hiển thị chậm hoặc xấu. Nên lọc nhóm khách hàng để xem chi tiết hơn.")

st.markdown("### 2. Bảng dữ liệu đại lý đã xử lý")
st.dataframe(df, use_container_width=True, hide_index=True)

st.markdown("### 3. Biểu đồ phân tích dữ liệu")
if chart_type == "Biểu đồ cột chồng":
    fig, ax = plt.subplots(figsize=(12,5))
    ind = np.arange(len(df))
    ax.bar(ind, df['Tổng bán trừ trả hàng'], width=0.5, label='Tổng bán cá nhân')
    ax.bar(ind, df['override_comm'], width=0.5, bottom=df['Tổng bán trừ trả hàng'], label='Hoa hồng hệ thống')
    ax.set_ylabel('Số tiền (VND)')
    ax.set_title('Tổng bán & Hoa hồng hệ thống từng cá nhân')
    ax.set_xticks(ind)
    ax.set_xticklabels(df['Tên khách hàng'], rotation=60, ha='right')
    ax.legend()
    st.pyplot(fig)
elif chart_type == "Sơ đồ Sunburst":
    try:
        fig2 = px.sunburst(
            df,
            path=['Nhóm khách hàng', 'Tên khách hàng'],
            values='Tổng bán trừ trả hàng',
            title="Sơ đồ hệ thống cấp bậc & doanh số"
        )
        st.plotly_chart(fig2, use_container_width=True)
    except Exception as e:
        st.error(f"Lỗi khi vẽ Sunburst chart: {e}")
elif chart_type == "Biểu đồ Pareto":
    try:
        df_sorted = df.sort_values('Tổng bán trừ trả hàng', ascending=False)
        cum_sum = df_sorted['Tổng bán trừ trả hàng'].cumsum()
        cum_perc = 100 * cum_sum / df_sorted['Tổng bán trừ trả hàng'].sum()
        fig3, ax1 = plt.subplots(figsize=(10,5))
        ax1.bar(np.arange(len(df_sorted)), df_sorted['Tổng bán trừ trả hàng'], label="Doanh số")
        ax1.set_ylabel('Doanh số')
        ax1.set_xticks(range(len(df_sorted)))
        ax1.set_xticklabels(df_sorted['Tên khách hàng'], rotation=60, ha='right')
        ax2 = ax1.twinx()
        ax2.plot(np.arange(len(df_sorted)), cum_perc, color='red', marker='o', label='Tích lũy (%)')
        ax2.set_ylabel('Tỷ lệ tích lũy (%)')
        ax1.set_title('Biểu đồ Pareto: Doanh số & tỷ trọng tích lũy')
        fig3.tight_layout()
        st.pyplot(fig3)
    except Exception as e:
        st.error(f"Lỗi khi vẽ Pareto chart: {e}")
elif chart_type == "Biểu đồ tròn (Pie)":
    try:
        fig4, ax4 = plt.subplots(figsize=(6,6))
        s = df.groupby('Nhóm khách hàng')['Tổng bán trừ trả hàng'].sum()
        ax4.pie(s, labels=s.index, autopct='%1.1f%%')
        ax4.set_title('Tỷ trọng doanh số theo nhóm khách hàng')
        st.pyplot(fig4)
    except Exception as e:
        st.error(f"Lỗi khi vẽ Pie chart: {e}")

st.markdown("### 4. Tải file kết quả định dạng màu nhóm vượt cấp")

output_file = f'sales_report_dep_{uuid.uuid4().hex[:6]}.xlsx'
df_export = df.sort_values(by=['parent_id', 'Mã khách hàng'], ascending=[True, True], na_position='last')
df_export.to_excel(output_file, index=False)

# ======= TÔ MÀU HỆ THỐNG VƯỢT CẤP (Trailblazer + Catalyst trực thuộc) =======
wb = load_workbook(output_file)
ws = wb.active
col_names = [cell.value for cell in ws[1]]
col_makh = col_names.index('Mã khách hàng')+1
col_vuotcap = col_names.index('vuot_cap_trailblazer')+1 if 'vuot_cap_trailblazer' in col_names else None

def pastel_color(seed_val):
    random.seed(str(seed_val))
    h = random.random()
    s = 0.28 + random.random()*0.09
    v = 0.97
    r, g, b = colorsys.hsv_to_rgb(h, s, v)
    return "%02X%02X%02X" % (int(r*255), int(g*255), int(b*255))

trailblazer_vuotcap = set(vuot_cap_ds.index)
trailblazer_to_color = {tb: PatternFill(start_color=pastel_color(tb+"vuotcap"), end_color=pastel_color(tb+"vuotcap"), fill_type='solid') for tb in trailblazer_vuotcap}

for row in range(2, ws.max_row + 1):
    ma_kh = str(ws.cell(row=row, column=col_makh).value)
    if ma_kh in trailblazer_to_color:
        fill = trailblazer_to_color[ma_kh]
    elif col_vuotcap and ws.cell(row=row, column=col_vuotcap).value in trailblazer_to_color:
        fill = trailblazer_to_color[ws.cell(row=row, column=col_vuotcap).value]
    else:
        fill = PatternFill(fill_type=None)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = fill

header_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
header_font = Font(bold=True, color='000000')
header_align = Alignment(horizontal='center', vertical='center')
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

bio = BytesIO()
try:
    wb.save(bio)
except PermissionError:
    st.error("Lỗi: File Excel đang mở ở chương trình khác. Đóng file lại trước khi export!")

downloaded = st.download_button(
    label="📥 Tải file Excel đã định dạng",
    data=bio.getvalue(),
    file_name=output_file,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
if downloaded:
    st.toast("✅ Đã tải xuống!", icon="✅")
